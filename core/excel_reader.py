"""Leitor de planilhas Excel para o Santa Rubi Label Studio.

Este módulo é responsável por:
- abrir arquivos .xlsx;
- ler a primeira aba da planilha;
- localizar colunas por nome de cabeçalho;
- validar a estrutura esperada;
- fornecer uma função para buscar produtos por código.

A lógica aqui é apenas de leitura e consulta. Nenhuma impressão ou geração
 de etiquetas é realizada.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ExcelReader:
    """Leitor simples de planilhas para consulta de produtos."""

    REQUIRED_COLUMNS = ["CODIGO", "CATEGORIA", "DESCRICAO", "PRECO"]
    HEADER_ALIASES = {
        "CODIGO": ["CODIGO", "CÓDIGO", "CÓDIGO SANTA RUBI"],
        "DESCRICAO": ["DESCRICAO", "DESCRIÇÃO", "DESCRIÇÃO PROD", "DESCRIÇÃO PROD (SISTEMA)"],
        "PRECO": ["PRECO", "PREÇO", "PREÇO DE VENDA"],
        "CATEGORIA": ["CATEGORIA"],
        "NUMERO": ["NUMERO", "Nº", "TAMANHO"],
        "QTD": ["QTD", "QUANTIDADE"],
    }

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)
        self.workbook = None
        self.sheet = None
        self.headers: list[str] = []
        self.rows: list[dict[str, Any]] = []
        self.total_products = 0
        self.column_map: dict[str, int] = {}
        self.error_message: str | None = None

        self._load()

    def _load(self) -> None:
        """Carrega a planilha e valida a estrutura esperada."""
        if not self.file_path.exists():
            self.error_message = "Arquivo não encontrado."
            return

        if self.file_path.suffix.lower() != ".xlsx":
            self.error_message = "Selecione um arquivo .xlsx."
            return

        try:
            self.workbook = load_workbook(filename=self.file_path, data_only=True, read_only=True)
        except Exception as exc:  # pragma: no cover - erro de leitura do Excel
            self.error_message = f"Não foi possível abrir a planilha: {exc}"
            return

        if not self.workbook.sheetnames:
            self.error_message = "A planilha não possui abas."
            self.workbook.close()
            self.workbook = None
            return

        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        rows = list(self.sheet.iter_rows(values_only=True))

        if not rows:
            self.error_message = "A planilha está vazia."
            self.workbook.close()
            self.workbook = None
            return

        self.headers = [self._normalize_header(value) for value in rows[0]]
        self._build_column_map()

        if self.error_message:
            self.workbook.close()
            self.workbook = None
            return

        for row_values in rows[1:]:
            if not row_values:
                continue

            row_dict = {}
            for column_name, index in self.column_map.items():
                row_dict[column_name] = row_values[index] if index < len(row_values) else None

            row_dict.setdefault("NUMERO", None)
            row_dict.setdefault("QTD", None)
            self.rows.append(row_dict)

        self.total_products = len(self.rows)

        if self.workbook is not None:
            self.workbook.close()
            self.workbook = None

    def _normalize_header(self, value: Any) -> str:
        """Normaliza o nome do cabeçalho para comparação robusta."""
        if value is None:
            return ""
        return str(value).strip().upper()

    def _build_column_map(self) -> None:
        """Cria um mapa entre os nomes esperados e as posições das colunas."""
        self.column_map = {}

        for index, header in enumerate(self.headers):
            normalized = self._normalize_header(header)
            for canonical_name, aliases in self.HEADER_ALIASES.items():
                if normalized in aliases:
                    self.column_map[canonical_name] = index
                    break

        missing_columns = [column for column in self.REQUIRED_COLUMNS if column not in self.column_map]
        if missing_columns:
            self.error_message = (
                "A planilha não contém todas as colunas obrigatórias. "
                f"Faltando: {', '.join(missing_columns)}"
            )
            return

        self.error_message = None

    def buscar_produto(self, codigo: str | int) -> dict[str, Any] | None:
        """Retorna um dicionário com os dados do produto encontrado."""
        if self.error_message is not None or self.rows is None:
            return None

        codigo_texto = str(codigo).strip()
        for row in self.rows:
            if str(row.get("CODIGO", "")).strip() == codigo_texto:
                return {
                    "codigo": row.get("CODIGO"),
                    "categoria": row.get("CATEGORIA"),
                    "descricao": row.get("DESCRICAO"),
                    "preco": row.get("PRECO"),
                    "numero": row.get("NUMERO"),
                    "qtd": row.get("QTD"),
                }

        return None
