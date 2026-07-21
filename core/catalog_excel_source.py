"""Fonte de dados do Catálogo Integrado a partir de um arquivo Excel.

Implementação concreta de DataSource. Deliberadamente independente do
ExcelReader usado pela impressão manual (aba existente) — as duas fontes de
dados não compartilham código, para que uma nunca afete o comportamento da
outra.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.catalog_datasource import DataSource


def normalize_header(value: Any) -> str:
    """Normaliza um cabeçalho para comparação: ignora caixa, espaços extras,
    quebras de linha e múltiplos espaços (mas preserva acentuação)."""
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().upper()


class ExcelCatalogSource(DataSource):
    """Lê abas, cabeçalhos e linhas diretamente de um arquivo .xlsx via openpyxl."""

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)

    def _open(self):
        if not self.file_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {self.file_path}")
        return load_workbook(filename=self.file_path, data_only=True, read_only=True)

    def list_sheets(self) -> list[str]:
        workbook = self._open()
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    def get_headers(self, sheet_names: list[str]) -> dict[str, list[str]]:
        workbook = self._open()
        try:
            headers: dict[str, list[str]] = {}
            for name in sheet_names:
                if name not in workbook.sheetnames:
                    headers[name] = []
                    continue
                sheet = workbook[name]
                first_row = next(sheet.iter_rows(max_row=1, values_only=True), ())
                headers[name] = [str(value) for value in first_row if value is not None]
            return headers
        finally:
            workbook.close()

    def count_rows(self, sheet_names: list[str]) -> dict[str, int]:
        workbook = self._open()
        try:
            counts: dict[str, int] = {}
            for name in sheet_names:
                if name not in workbook.sheetnames:
                    counts[name] = 0
                    continue
                sheet = workbook[name]
                counts[name] = sum(
                    1
                    for row in sheet.iter_rows(min_row=2, values_only=True)
                    if any(value is not None for value in row)
                )
            return counts
        finally:
            workbook.close()

    def read_rows(self, sheet_names: list[str]) -> list[dict[str, Any]]:
        workbook = self._open()
        try:
            rows: list[dict[str, Any]] = []
            for name in sheet_names:
                if name not in workbook.sheetnames:
                    continue
                sheet = workbook[name]
                sheet_rows = sheet.iter_rows(values_only=True)
                header_row = next(sheet_rows, None)
                if header_row is None:
                    continue
                headers = [str(value) if value is not None else "" for value in header_row]
                for row_values in sheet_rows:
                    if not any(value is not None for value in row_values):
                        continue
                    row_dict: dict[str, Any] = {
                        headers[index]: row_values[index]
                        for index in range(min(len(headers), len(row_values)))
                    }
                    row_dict["_sheet"] = name
                    rows.append(row_dict)
            return rows
        finally:
            workbook.close()
