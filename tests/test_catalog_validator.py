import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from core.catalog_excel_source import ExcelCatalogSource
from core.catalog_settings import CatalogSettings
from core.catalog_validator import CatalogConfigurationValidator


def build_workbook(temp_dir: str) -> Path:
    file_path = Path(temp_dir) / "catalogo.xlsx"
    workbook = Workbook()

    sheet1 = workbook.active
    sheet1.title = "DIP-15"
    sheet1.append(["CÓDIGO SANTA RUBI", "DESCRIÇÃO PROD (SISTEMA)", "PREÇO DE VENDA FINAL", "CATEGORIA", "TAMANHO Nº", "QTD"])
    sheet1.append(["1001", "Anel Prata", 59.9, "Aneis", "16", 5])
    sheet1.append(["1002", "Brinco Ouro", 89.9, "Brincos", None, 2])

    sheet2 = workbook.create_sheet("IW-13")
    sheet2.append(["CÓDIGO SANTA RUBI", "DESCRIÇÃO PROD (SISTEMA)", "PREÇO DE VENDA FINAL", "CATEGORIA", "TAMANHO Nº", "QTD"])
    sheet2.append(["2001", "Colar Ouro", 129.9, "Colares", None, 1])

    sheet3 = workbook.create_sheet("TESTES")
    sheet3.append(["CODIGO", "DESCRICAO"])  # sem as colunas Preco/Categoria/Numeracao
    sheet3.append(["9999", "Produto de teste"])

    workbook.save(file_path)
    return file_path


def base_settings(sheets: list[str]) -> CatalogSettings:
    return CatalogSettings(
        file_path="",
        selected_sheets=sheets,
        column_map={
            "codigo": "CÓDIGO SANTA RUBI",
            "descricao": "DESCRIÇÃO PROD (SISTEMA)",
            "preco": "PREÇO DE VENDA FINAL",
            "categoria": "CATEGORIA",
            "numeracao": "TAMANHO Nº",
            "quantidade": "QTD",
        },
    )


class CatalogConfigurationValidatorTests(unittest.TestCase):
    def test_validate_happy_path(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = build_workbook(temp_dir)
            settings = base_settings(["DIP-15", "IW-13"])
            settings.file_path = str(file_path)
            validator = CatalogConfigurationValidator(ExcelCatalogSource(file_path))

            report = validator.validate(settings)

            self.assertTrue(report.file_found)
            self.assertEqual(report.sheets_found, ["DIP-15", "IW-13"])
            self.assertTrue(report.all_columns_found)
            self.assertEqual(report.total_products, 3)
            self.assertEqual(report.issues, [])
            self.assertTrue(report.is_valid)

    def test_validate_reports_missing_sheet(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = build_workbook(temp_dir)
            settings = base_settings(["DIP-15", "NAO_EXISTE"])
            validator = CatalogConfigurationValidator(ExcelCatalogSource(file_path))

            report = validator.validate(settings)

            self.assertTrue(any("NAO_EXISTE" in issue.message for issue in report.issues))
            self.assertEqual(report.sheets_found, ["DIP-15"])

    def test_validate_reports_missing_column_with_exact_sheet_and_field(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = build_workbook(temp_dir)
            settings = base_settings(["DIP-15", "TESTES"])
            validator = CatalogConfigurationValidator(ExcelCatalogSource(file_path))

            report = validator.validate(settings)

            self.assertFalse(report.all_columns_found)
            self.assertFalse(report.is_valid)
            preco_issues = [i for i in report.issues if i.field == "preco" and i.sheet == "TESTES"]
            self.assertEqual(len(preco_issues), 1)
            self.assertIn("TESTES", preco_issues[0].message)
            self.assertIn("PREÇO DE VENDA FINAL", preco_issues[0].message)

    def test_validate_reports_unmapped_field(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = build_workbook(temp_dir)
            settings = base_settings(["DIP-15"])
            settings.column_map.pop("preco")
            validator = CatalogConfigurationValidator(ExcelCatalogSource(file_path))

            report = validator.validate(settings)

            unmapped_issues = [i for i in report.issues if i.field == "preco" and i.sheet is None]
            self.assertEqual(len(unmapped_issues), 1)
            self.assertFalse(report.all_columns_found)

    def test_validate_missing_file(self):
        settings = base_settings(["DIP-15"])
        validator = CatalogConfigurationValidator(ExcelCatalogSource("arquivo_inexistente.xlsx"))

        report = validator.validate(settings)

        self.assertFalse(report.file_found)
        self.assertFalse(report.is_valid)


if __name__ == "__main__":
    unittest.main()
