import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from core.excel_reader import ExcelReader


class ExcelReaderTests(unittest.TestCase):
    def test_load_workbook_and_find_product(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "produtos.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Planilha1"
            sheet.append(["CODIGO", "CATEGORIA", "DESCRICAO", "PRECO", "NUMERO"])
            sheet.append(["1001", "Bebidas", "Coca-Cola", 4.5, 10])
            workbook.save(file_path)

            reader = ExcelReader(file_path)
            self.assertEqual(reader.total_products, 1)

            product = reader.buscar_produto("1001")
            self.assertIsNotNone(product)
            self.assertEqual(product["codigo"], "1001")
            self.assertEqual(product["categoria"], "Bebidas")
            self.assertEqual(product["descricao"], "Coca-Cola")
            self.assertEqual(product["preco"], 4.5)
            self.assertEqual(product["numero"], 10)

    def test_load_workbook_with_santa_rubi_headers(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "santarubi.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Planilha1"
            sheet.append(["CÓDIGO SANTA RUBI", "CATEGORIA", "DESCRIÇÃO PROD (SISTEMA)", "PREÇO DE VENDA", "QTD"])
            sheet.append(["2001", "Lanches", "Hambúrguer", 12.99, 3])
            workbook.save(file_path)

            reader = ExcelReader(file_path)
            self.assertEqual(reader.total_products, 1)

            product = reader.buscar_produto("2001")
            self.assertIsNotNone(product)
            self.assertEqual(product["codigo"], "2001")
            self.assertEqual(product["categoria"], "Lanches")
            self.assertEqual(product["descricao"], "Hambúrguer")
            self.assertEqual(product["preco"], 12.99)
            self.assertEqual(product["qtd"], 3)


if __name__ == "__main__":
    unittest.main()
