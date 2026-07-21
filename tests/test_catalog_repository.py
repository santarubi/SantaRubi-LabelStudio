import tempfile
import unittest
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from core.catalog_datasource import DataSource
from core.catalog_excel_source import ExcelCatalogSource
from core.catalog_repository import CatalogRepository
from core.catalog_settings import CatalogSettings
from core.print_item import normalize_quantity


def build_workbook(temp_dir: str) -> Path:
    file_path = Path(temp_dir) / "catalogo.xlsx"
    workbook = Workbook()

    sheet1 = workbook.active
    sheet1.title = "DIP-15"
    sheet1.append(["CÓDIGO SANTA RUBI", "DESCRIÇÃO PROD (SISTEMA)", "PREÇO DE VENDA FINAL", "CATEGORIA", "TAMANHO Nº"])
    sheet1.append(["1001", "Anel Prata", 59.9, "Aneis", "16"])
    sheet1.append(["1002", "Brinco Ouro", 89.9, "Brincos", None])

    sheet2 = workbook.create_sheet("IW-13")
    sheet2.append(["CÓDIGO SANTA RUBI", "DESCRIÇÃO PROD (SISTEMA)", "PREÇO DE VENDA FINAL", "CATEGORIA", "TAMANHO Nº"])
    sheet2.append(["2001", "Colar Ouro", 129.9, "Colares", None])

    workbook.save(file_path)
    return file_path


def base_settings(sheets: list[str]) -> CatalogSettings:
    return CatalogSettings(
        selected_sheets=sheets,
        column_map={
            "codigo": "CÓDIGO SANTA RUBI",
            "descricao": "DESCRIÇÃO PROD (SISTEMA)",
            "preco": "PREÇO DE VENDA FINAL",
            "categoria": "CATEGORIA",
            "numeracao": "TAMANHO Nº",
        },
    )


class FakeDataSource(DataSource):
    """DataSource em memória, só para contar quantas vezes cada operação é
    realmente chamada — usada para provar que o Repository não relê a
    origem em toda pesquisa/filtro, só em load()/reload()."""

    def __init__(self, sheets_data: dict[str, list[dict[str, Any]]]):
        self.sheets_data = sheets_data
        self.read_rows_calls = 0
        self.list_sheets_calls = 0

    def list_sheets(self) -> list[str]:
        self.list_sheets_calls += 1
        return list(self.sheets_data.keys())

    def get_headers(self, sheet_names: list[str]) -> dict[str, list[str]]:
        headers: dict[str, list[str]] = {}
        for name in sheet_names:
            rows = self.sheets_data.get(name, [])
            headers[name] = list(rows[0].keys()) if rows else []
        return headers

    def count_rows(self, sheet_names: list[str]) -> dict[str, int]:
        return {name: len(self.sheets_data.get(name, [])) for name in sheet_names}

    def read_rows(self, sheet_names: list[str]) -> list[dict[str, Any]]:
        self.read_rows_calls += 1
        rows: list[dict[str, Any]] = []
        for name in sheet_names:
            for row in self.sheets_data.get(name, []):
                row_with_sheet = dict(row)
                row_with_sheet["_sheet"] = name
                rows.append(row_with_sheet)
        return rows


class CatalogRepositoryHeadersTests(unittest.TestCase):
    def test_get_available_headers_dedups_and_keeps_first_spelling(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = build_workbook(temp_dir)
            repository = CatalogRepository(ExcelCatalogSource(file_path))
            headers = repository.get_available_headers(["DIP-15", "IW-13"])
            self.assertEqual(
                headers,
                ["CÓDIGO SANTA RUBI", "DESCRIÇÃO PROD (SISTEMA)", "PREÇO DE VENDA FINAL", "CATEGORIA", "TAMANHO Nº"],
            )


class CatalogRepositoryLoadTests(unittest.TestCase):
    def test_load_converts_rows_into_catalog_products(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = build_workbook(temp_dir)
            settings = base_settings(["DIP-15", "IW-13"])
            repository = CatalogRepository(ExcelCatalogSource(file_path))

            products = repository.load(settings)

            self.assertEqual(len(products), 3)
            first = next(p for p in products if p.codigo == "1001")
            self.assertEqual(first.descricao, "Anel Prata")
            self.assertEqual(first.preco, 59.9)
            self.assertEqual(first.categoria, "Aneis")
            self.assertEqual(first.numeracao, "16")
            self.assertEqual(first.fornecedor, "DIP-15")

            third = next(p for p in products if p.codigo == "2001")
            self.assertEqual(third.fornecedor, "IW-13")

    def test_products_before_load_raises(self):
        repository = CatalogRepository(ExcelCatalogSource("qualquer.xlsx"))
        with self.assertRaises(RuntimeError):
            _ = repository.products

    def test_is_loaded_flag(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = build_workbook(temp_dir)
            settings = base_settings(["DIP-15"])
            repository = CatalogRepository(ExcelCatalogSource(file_path))
            self.assertFalse(repository.is_loaded)
            repository.load(settings)
            self.assertTrue(repository.is_loaded)


class CatalogRepositoryQuantityTests(unittest.TestCase):
    """A quantidade padrão da coluna QTD NÃO é atribuída a CatalogProduct —
    o Repository guarda o valor bruto (sem normalizar) em
    `attributes["default_quantity"]`; a normalização (vazio/inválido -> 1,
    decimal -> inteiro) é responsabilidade de PrintItem/normalize_quantity,
    aplicada só quando uma solicitação de impressão é de fato criada."""

    def _build_workbook_with_quantity(self, temp_dir: str) -> Path:
        file_path = Path(temp_dir) / "catalogo_qtd.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "DIP-15"
        sheet.append(["CODIGO", "DESCRICAO", "PRECO", "CATEGORIA", "NUMERO", "QTD"])
        sheet.append(["1001", "Anel Prata", 59.9, "Aneis", "16", 5])
        sheet.append(["1002", "Brinco Ouro", 89.9, "Brincos", None, None])  # vazia
        sheet.append(["1003", "Colar Ouro", 129.9, "Colares", None, "abc"])  # invalida
        sheet.append(["1004", "Pulseira Prata", 49.9, "Pulseiras", None, 3.7])  # decimal
        workbook.save(file_path)
        return file_path

    def _settings_with_quantity(self) -> CatalogSettings:
        return CatalogSettings(
            selected_sheets=["DIP-15"],
            column_map={
                "codigo": "CODIGO",
                "descricao": "DESCRICAO",
                "preco": "PRECO",
                "categoria": "CATEGORIA",
                "numeracao": "NUMERO",
                "quantidade": "QTD",
            },
        )

    def test_repository_stores_raw_quantity_in_attributes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = self._build_workbook_with_quantity(temp_dir)
            settings = self._settings_with_quantity()
            repository = CatalogRepository(ExcelCatalogSource(file_path))

            products = repository.load(settings)
            by_codigo = {p.codigo: p for p in products}

            # Repository guarda o valor BRUTO (sem normalizar) em attributes.
            self.assertEqual(by_codigo["1001"].attributes["default_quantity"], 5)
            self.assertIsNone(by_codigo["1002"].attributes["default_quantity"])
            self.assertEqual(by_codigo["1003"].attributes["default_quantity"], "abc")
            self.assertEqual(by_codigo["1004"].attributes["default_quantity"], 3.7)

            # normalize_quantity() e' quem aplica as regras de negocio sobre
            # esse valor bruto (mesma funcao usada por CatalogService/PrintItem).
            self.assertEqual(normalize_quantity(by_codigo["1001"].attributes["default_quantity"]), 5)
            self.assertEqual(normalize_quantity(by_codigo["1002"].attributes["default_quantity"]), 1)
            self.assertEqual(normalize_quantity(by_codigo["1003"].attributes["default_quantity"]), 1)
            self.assertEqual(normalize_quantity(by_codigo["1004"].attributes["default_quantity"]), 3)

    def test_repository_leaves_default_quantity_none_when_column_not_mapped(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = self._build_workbook_with_quantity(temp_dir)
            settings = self._settings_with_quantity()
            settings.column_map.pop("quantidade")
            repository = CatalogRepository(ExcelCatalogSource(file_path))

            products = repository.load(settings)
            self.assertTrue(all(p.attributes.get("default_quantity") is None for p in products))
            self.assertTrue(all(normalize_quantity(p.attributes.get("default_quantity")) == 1 for p in products))

    def test_products_do_not_expose_quantidade_field(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = self._build_workbook_with_quantity(temp_dir)
            settings = self._settings_with_quantity()
            repository = CatalogRepository(ExcelCatalogSource(file_path))

            products = repository.load(settings)
            self.assertTrue(all(not hasattr(p, "quantidade") for p in products))


class CatalogRepositoryCacheTests(unittest.TestCase):
    """Prova que, depois de load(), nenhuma leitura adicional acontece na
    DataSource — pesquisa/filtro/acesso repetido a `products` usam só a
    lista em memória."""

    def _fake_source(self) -> FakeDataSource:
        return FakeDataSource(
            {
                "DIP-15": [
                    {"CÓDIGO SANTA RUBI": "1001", "DESCRIÇÃO PROD (SISTEMA)": "Anel Prata",
                     "PREÇO DE VENDA FINAL": 59.9, "CATEGORIA": "Aneis", "TAMANHO Nº": "16"},
                    {"CÓDIGO SANTA RUBI": "1002", "DESCRIÇÃO PROD (SISTEMA)": "Brinco Ouro",
                     "PREÇO DE VENDA FINAL": 89.9, "CATEGORIA": "Brincos", "TAMANHO Nº": None},
                ]
            }
        )

    def test_products_access_does_not_trigger_new_read(self):
        source = self._fake_source()
        settings = base_settings(["DIP-15"])
        repository = CatalogRepository(source)

        repository.load(settings)
        self.assertEqual(source.read_rows_calls, 1)

        # simula "pesquisa" e "filtro": varios acessos a .products
        for _ in range(5):
            _ = [p for p in repository.products if "anel" in p.search_blob]

        self.assertEqual(source.read_rows_calls, 1, "acessar products/pesquisar nao deveria reler a fonte")

    def test_reload_triggers_new_read(self):
        source = self._fake_source()
        settings = base_settings(["DIP-15"])
        repository = CatalogRepository(source)

        repository.load(settings)
        self.assertEqual(source.read_rows_calls, 1)

        repository.reload(settings)
        self.assertEqual(source.read_rows_calls, 2, "reload() deve reler a fonte de propósito")

    def test_reload_replaces_cache_with_fresh_data(self):
        source = self._fake_source()
        settings = base_settings(["DIP-15"])
        repository = CatalogRepository(source)
        repository.load(settings)
        self.assertEqual(len(repository.products), 2)

        # simula planilha atualizada com um produto a mais
        source.sheets_data["DIP-15"].append(
            {"CÓDIGO SANTA RUBI": "1003", "DESCRIÇÃO PROD (SISTEMA)": "Colar Prata",
             "PREÇO DE VENDA FINAL": 149.9, "CATEGORIA": "Colares", "TAMANHO Nº": None}
        )
        repository.reload(settings)
        self.assertEqual(len(repository.products), 3)


if __name__ == "__main__":
    unittest.main()
